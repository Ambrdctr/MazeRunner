
# coding: utf-8

# # Imports

# In[183]:

from random import *
from openpyxl.styles.borders import Border, Side
from openpyxl import Workbook
import time


# # Class

# ## Cell class

# In[184]:

class Cell:

    #Constructor
    def __init__(self, i):
        self.walls = [True, True, True, True]
        self.state = 'v'
        self.set = i


# ## Grid class

# In[185]:

class Grid:

    def __init__(self, h, w):
        self.w = w
        self.h = h
        grid = []
        ind = 0
        for i in range(w):
            grid.append([])
            for j in range(h):
                # Ajout de l'indice pour le calcul des ensemble
                initial_cell = Cell(ind)
                grid[i].append(initial_cell)
                ind += 1
        self.grid = grid

    def is_cell(self,c,wall):
        if wall == 0:
            if c[0] == 0:
                return False
        elif wall == 1:
            if c[1] == self.h-1:
                return False
        elif wall == 2:
            if c[0] == self.w-1:
                return False
        elif wall == 3:
            if c[1] == 0:
                return False
        return True

    def remove_wall(self,c1, c2, w):
        #print(wall, (wall+2)%4)
        #print(c1,wall,c2,(wall+2)%4)
        self.grid[c1[0]][c1[1]].walls[w] = False
        self.grid[c2[0]][c2[1]].walls[(w+2)%4] = False


# ## Set class

# In[186]:

class Set:

    def __init__(self):
        self.lst = {}

    def already(self, ind, c):
        if ind in self.lst:
            if c in self.lst[ind]:
                return True
        return False

    def add_set(self, ind, c):
        if not self.already(ind, c):
            self.lst[ind] = [c]

    def same_set(self, c1, c2, grid):
        idc1 = grid.grid[c1[0]][c1[1]].set
        idc2 = grid.grid[c2[0]][c2[1]].set
        if idc1 < idc2:
            for e in self.lst[idc2]:
                grid.grid[e[0]][e[1]].set = idc1
            self.lst[idc1] += self.lst[idc2]
            del self.lst[idc2]
        else:
            for e in self.lst[idc1]:
                grid.grid[e[0]][e[1]].set = idc2
            self.lst[idc2]+= self.lst[idc1]
            del self.lst[idc1]

    def different_sets(self, c1, c2, grid):
        return grid.grid[c1[0]][c1[1]].set != grid.grid[c2[0]][c2[1]].set

    def all_same_set(self, grid, size):
        return len(self.lst) == 1 and next(iter(self.lst)) == 0 and size == len(self.lst[0])


# # Principal program

# In[187]:

def choseAndRemove(set, grid):
    size = grid.w * grid.h
    found = False
    while not set.all_same_set(grid, size):
        # Choix d'un mur aleatoire
        x = randint(0, grid.w - 1)
        y = randint(0, grid.h - 1)
        wall = randint(0,3)
        c1 = (x, y)
        #print(c1, wall)
        set.add_set(grid.grid[c1[0]][c1[1]].set, c1)
        # Case derriere le mur
        if grid.is_cell(c1, wall):
            if grid.grid[c1[0]][c1[1]].walls[wall] == True:
                if wall == 0 and set.different_sets(c1, (x-1,y), grid): #top
                    c2 = (x-1,y)
                    found = True
                elif wall == 1 and set.different_sets(c1, (x,y+1), grid): #right
                    c2 = (x,y+1)
                    found = True
                elif wall == 2 and set.different_sets(c1, (x+1,y), grid): #bottom
                    c2 = (x+1,y)
                    found = True
                elif wall == 3 and set.different_sets(c1, (x,y-1), grid): #left
                    c2 = (x,y-1)
                    found = True
                if found:
                    set.add_set(grid.grid[c2[0]][c2[1]].set, c2)
                    set.same_set(c1,c2,grid)
                    grid.remove_wall(c1,c2,wall)
                    found = False
                    #print(c2)
        #print(set.lst)
        #disp_maze(grid)
    #print(len(set.lst[0]))
    #print(set.lst)
    #disp_maze(grid)
        

def kruskal(maze):
    cellSet = Set()
    choseAndRemove(cellSet,maze)
    return maze

def disp_maze(maze):
    chaine = ""
    for i in range(maze.w):
        for j in range(maze.h):
            chaine += str(maze.grid[i][j].walls) + " " + str(maze.grid[i][j].set)
        chaine += "\n"
    return print(chaine)

def colnum_string(n):
    string = ""
    while n > 0:
        n, remainder = divmod(n - 1, 26)
        string = chr(65 + remainder) + string
    return string

def maze_to_xlsx(maze, file):
    wb = Workbook()
    ws = wb.active
    for i in range(maze.w):
        for j in range(maze.h):
            top = Side(color="FFFFFFFF")
            right = Side(color="FFFFFFFF")
            bottom = Side(color="FFFFFFFF")
            left = Side(color="FFFFFFFF")
            for k in range(4):
                if maze.grid[i][j].walls[k]:
                    if k == 0:
                        top = Side(style='thick')
                    if k == 1:
                        right = Side(style='thick')
                    if k == 2:
                        bottom = Side(style='thick')
                    if k == 3:
                        left = Side(style='thick')
            border = Border(left, right, top, bottom) # Attention à l'ordre !
            ws.cell(row=i+2, column=j+2).border = border
            ws.cell(row=1, column=j+2).value = j
            ws.cell(row=i+2, column=1).value = i
    for i in range(2, maze.h+2):
        ws.column_dimensions[colnum_string(i)].width = 3
    wb.save(file)
    


def create_maze(w,h,file):
    debut = time.time()
    maze = Grid(w,h)
    #disp_maze(maze)
    #maze_to_xlsx(maze, file)
    kruskal(maze)
    #disp_maze(maze)
    maze_to_xlsx(maze, file)
    fin = time.time()
    print('Fini. ' + file + ' généré en : ' + str(round(fin-debut,3)) + ' secondes')
    return maze


# # Run

# In[188]:

#create_maze(1,2,'maze1x2.xlsx')
#create_maze(2,2,'maze2x2.xlsx')
#create_maze(3,3,'maze3x3.xlsx')
#create_maze(5,5,'maze5x5.xlsx')
#create_maze(10,10,'maze10x10.xlsx')
#create_maze(22,15,'maze22x15.xlsx')
create_maze(30,20,'maze30x20.xlsx')
#create_maze(50,50,'maze50x50.xlsx')
#create_maze(100,100,'maze100x100.xlsx')


# In[ ]:



